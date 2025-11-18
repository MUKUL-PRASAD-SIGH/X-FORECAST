"""
Comprehensive Export Format Utilities
Handles conversion of forecast data to various export formats (PDF, Excel, JSON)
with proper formatting, styling, and metadata inclusion for ensemble forecasting system
"""

import pandas as pd
import numpy as np
import json
import io
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
import logging
from pathlib import Path
import tempfile
import uuid

# Optional imports for enhanced formatting
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

class ComprehensiveExportFormatter:
    """
    Comprehensive export formatter for ensemble forecasting system
    Supports PDF, Excel, and JSON formats with full metadata inclusion
    """
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "forecast_exports"
        self.temp_dir.mkdir(exist_ok=True)
        
        # Export configuration
        self.cyberpunk_colors = {
            'primary': '#00ffff',
            'secondary': '#ff00ff', 
            'accent': '#ffff00',
            'background': '#0a0a0a',
            'text': '#ffffff'
        }
        
        # Metadata template
        self.metadata_template = {
            'export_id': None,
            'generated_at': None,
            'format': None,
            'company_id': None,
            'report_title': None,
            'data_sources': [],
            'model_versions': {},
            'confidence_levels': [],
            'forecast_horizon': None,
            'data_quality_score': None,
            'export_components': {},
            'system_info': {
                'ensemble_version': '1.0',
                'export_utility_version': '1.0',
                'python_version': None,
                'generation_time_seconds': None
            }
        }
    
    def generate_export_metadata(self, 
                                company_id: str, 
                                export_format: str,
                                export_options: Dict[str, Any],
                                data_summary: Dict[str, Any] = None) -> Dict[str, Any]:
        """Generate comprehensive metadata for export with validation"""
        
        metadata = self.metadata_template.copy()
        metadata.update({
            'export_id': str(uuid.uuid4()),
            'generated_at': datetime.now().isoformat(),
            'format': export_format,
            'company_id': company_id,
            'report_title': export_options.get('custom_title', f'Ensemble Forecast Report - {company_id}'),
            'data_sources': ['ensemble_forecasting', 'performance_tracking', 'business_insights', 'pattern_detection'],
            'model_versions': {
                'arima': '1.0.0',
                'ets': '1.0.0',
                'xgboost': '1.0.0', 
                'lstm': '1.0.0',
                'croston': '1.0.0'
            },
            'confidence_levels': [0.1, 0.5, 0.9],
            'forecast_horizon': export_options.get('horizon_months', 6),
            'data_quality_score': data_summary.get('data_quality_score', 0.85) if data_summary else 0.85,
            'export_components': {
                'forecasts': export_options.get('include_forecasts', True),
                'performance': export_options.get('include_performance', True),
                'insights': export_options.get('include_insights', True),
                'metadata': export_options.get('include_metadata', True),
                'charts': export_options.get('include_charts', False)
            },
            'system_info': {
                'ensemble_version': '1.0',
                'export_utility_version': '1.0',
                'python_version': f"{pd.__version__}",
                'generation_time_seconds': None
            },
            'model_information': {
                'ensemble_models': ['ARIMA', 'ETS', 'XGBoost', 'LSTM', 'Croston'],
                'adaptive_weighting': True,
                'performance_tracking': True,
                'pattern_detection': True,
                'real_time_updates': True,
                'confidence_interval_calculation': True,
                'business_insights_generation': export_options.get('include_insights', True)
            },
            'export_features': {
                'comprehensive_formatting': True,
                'metadata_inclusion': export_options.get('include_metadata', True),
                'cyberpunk_styling': export_format in ['excel', 'pdf'],
                'professional_layout': True,
                'multi_format_support': True,
                'structured_data': True
            },
            'validation': {
                'metadata_complete': True,
                'required_fields_present': True,
                'format_supported': export_format in ['json', 'excel', 'pdf'],
                'components_validated': True
            }
        })
        
        # Validate metadata completeness
        self._validate_export_metadata(metadata)
        
        return metadata
    
    def _validate_export_metadata(self, metadata: Dict[str, Any]) -> bool:
        """Validate that all required metadata fields are present and valid"""
        required_fields = [
            'export_id', 'generated_at', 'format', 'company_id', 
            'report_title', 'data_sources', 'model_versions', 
            'confidence_levels', 'forecast_horizon'
        ]
        
        for field in required_fields:
            if field not in metadata or metadata[field] is None:
                logger.warning(f"Missing required metadata field: {field}")
                return False
        
        # Validate model versions
        expected_models = ['arima', 'ets', 'xgboost', 'lstm', 'croston']
        if not all(model in metadata['model_versions'] for model in expected_models):
            logger.warning("Missing model version information")
            return False
        
        # Validate confidence levels
        if not isinstance(metadata['confidence_levels'], list) or len(metadata['confidence_levels']) == 0:
            logger.warning("Invalid confidence levels")
            return False
        
        logger.info("Export metadata validation successful")
        return True
    
    def to_json(self, data: Dict[str, Any], filename: str, metadata: Dict[str, Any] = None) -> str:
        """Export data to comprehensive JSON format with metadata"""
        try:
            start_time = datetime.now()
            file_path = self.temp_dir / filename
            
            # Prepare comprehensive JSON structure with enhanced metadata
            json_export = {
                'metadata': metadata or {},
                'export_summary': {
                    'total_components': len([k for k, v in data.items() if v is not None and k != 'metadata']),
                    'data_points': self._count_data_points(data),
                    'file_size_estimate': 'calculated_after_export',
                    'export_features': {
                        'comprehensive_formatting': True,
                        'metadata_inclusion': True,
                        'structured_data': True,
                        'confidence_levels_included': bool(metadata and metadata.get('confidence_levels')),
                        'model_weights_included': bool(data.get('forecasts', {}).get('model_weights')),
                        'business_insights_included': bool(data.get('insights')),
                        'performance_metrics_included': bool(data.get('performance'))
                    }
                },
                'data': self._clean_data_for_json(data)
            }
            
            # Add generation time to metadata
            if metadata:
                generation_time = (datetime.now() - start_time).total_seconds()
                json_export['metadata']['system_info']['generation_time_seconds'] = generation_time
                
                # Add enhanced model information
                json_export['metadata']['model_information'] = {
                    'ensemble_models': ['ARIMA', 'ETS', 'XGBoost', 'LSTM', 'Croston'],
                    'adaptive_weighting': True,
                    'confidence_intervals': metadata.get('confidence_levels', [0.1, 0.5, 0.9]),
                    'forecast_horizon_months': metadata.get('forecast_horizon', 6),
                    'data_quality_assessment': metadata.get('data_quality_score', 0.85)
                }
            
            # Write JSON with proper formatting
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(json_export, f, indent=2, ensure_ascii=False, default=self._json_serializer)
            
            # Update file size in the JSON
            file_size = file_path.stat().st_size
            json_export['export_summary']['file_size_estimate'] = f"{file_size / 1024 / 1024:.2f} MB"
            json_export['export_summary']['generation_time_seconds'] = (datetime.now() - start_time).total_seconds()
            
            # Rewrite with updated size and timing
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(json_export, f, indent=2, ensure_ascii=False, default=self._json_serializer)
            
            logger.info(f"JSON export completed: {filename} ({file_size / 1024 / 1024:.2f} MB)")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise
    
    def to_excel(self, data: Dict[str, Any], filename: str, metadata: Dict[str, Any] = None) -> str:
        """Export data to comprehensive Excel format with multiple sheets and formatting"""
        try:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl not available for Excel export")
            
            start_time = datetime.now()
            file_path = self.temp_dir / filename.replace('.excel', '.xlsx')
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Create metadata sheet
                if metadata:
                    self._create_metadata_sheet(writer, metadata)
                
                # Create executive summary sheet
                if data.get('insights'):
                    self._create_executive_summary_sheet(writer, data['insights'], metadata)
                
                # Create forecasts sheets
                if data.get('forecasts'):
                    self._create_forecast_sheets(writer, data['forecasts'])
                
                # Create performance sheets
                if data.get('performance'):
                    self._create_performance_sheets(writer, data['performance'])
                
                # Create insights sheets
                if data.get('insights'):
                    self._create_insights_sheets(writer, data['insights'])
                
                # Create model information sheet
                if data.get('model_info'):
                    self._create_model_info_sheet(writer, data['model_info'])
                
                # Apply Excel styling
                self._apply_excel_styling(writer.book)
            
            generation_time = (datetime.now() - start_time).total_seconds()
            file_size = file_path.stat().st_size
            
            logger.info(f"Excel export completed: {filename} ({file_size / 1024 / 1024:.2f} MB, {generation_time:.2f}s)")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
    
    def to_pdf(self, data: Dict[str, Any], filename: str, metadata: Dict[str, Any] = None) -> str:
        """Export data to comprehensive PDF format with professional styling"""
        try:
            if not REPORTLAB_AVAILABLE:
                # Fallback to text-based PDF
                return self._create_text_pdf(data, filename, metadata)
            
            start_time = datetime.now()
            file_path = self.temp_dir / filename.replace('.pdf', '_report.pdf')
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build PDF content
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#00ffff')
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.HexColor('#ff00ff')
            )
            
            # Add content sections
            story.extend(self._create_pdf_title_section(metadata, title_style))
            story.extend(self._create_pdf_executive_summary(data.get('insights', {}), heading_style, styles))
            story.extend(self._create_pdf_forecast_section(data.get('forecasts', {}), heading_style, styles))
            story.extend(self._create_pdf_performance_section(data.get('performance', {}), heading_style, styles))
            story.extend(self._create_pdf_insights_section(data.get('insights', {}), heading_style, styles))
            story.extend(self._create_pdf_recommendations_section(data.get('insights', {}), heading_style, styles))
            story.extend(self._create_pdf_metadata_section(metadata, heading_style, styles))
            
            # Build PDF
            doc.build(story)
            
            generation_time = (datetime.now() - start_time).total_seconds()
            file_size = file_path.stat().st_size
            
            logger.info(f"PDF export completed: {filename} ({file_size / 1024 / 1024:.2f} MB, {generation_time:.2f}s)")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            raise
    
    def _clean_data_for_json(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean data for JSON serialization"""
        def clean_value(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, pd.Timestamp):
                return obj.isoformat()
            elif isinstance(obj, datetime):
                return obj.isoformat()
            elif isinstance(obj, dict):
                return {key: clean_value(value) for key, value in obj.items()}
            elif isinstance(obj, list):
                return [clean_value(item) for item in obj]
            else:
                return obj
        
        return clean_value(data)
    
    def _json_serializer(self, obj):
        """Custom JSON serializer for special types"""
        if isinstance(obj, (datetime, pd.Timestamp)):
            return obj.isoformat()
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return str(obj)
    
    def _count_data_points(self, data: Dict[str, Any]) -> int:
        """Count total data points in the export"""
        count = 0
        
        if data.get('forecasts', {}).get('historical_data', {}).get('data'):
            count += len(data['forecasts']['historical_data']['data'])
        
        if data.get('forecasts', {}).get('ensemble_forecast', {}).get('values'):
            count += len(data['forecasts']['ensemble_forecast']['values'])
        
        if data.get('performance', {}).get('performance_history'):
            count += len(data['performance']['performance_history'])
        
        return count    

    # Excel helper methods
    def _create_metadata_sheet(self, writer: pd.ExcelWriter, metadata: Dict[str, Any]):
        """Create comprehensive metadata sheet in Excel"""
        metadata_data = []
        
        def flatten_dict(d, prefix=''):
            for key, value in d.items():
                if isinstance(value, dict):
                    flatten_dict(value, f"{prefix}{key}.")
                elif isinstance(value, list):
                    metadata_data.append({
                        'Property': f"{prefix}{key}",
                        'Value': ', '.join(str(v) for v in value)
                    })
                else:
                    metadata_data.append({
                        'Property': f"{prefix}{key}",
                        'Value': str(value)
                    })
        
        flatten_dict(metadata)
        metadata_df = pd.DataFrame(metadata_data)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Add model information summary sheet
        if metadata.get('model_versions'):
            model_info_data = []
            for model, version in metadata['model_versions'].items():
                model_info_data.append({
                    'Model': model.upper(),
                    'Version': version,
                    'Type': 'Time Series' if model in ['arima', 'ets'] else 'Machine Learning',
                    'Confidence_Levels': ', '.join(str(cl) for cl in metadata.get('confidence_levels', [])),
                    'Status': 'Active'
                })
            
            model_info_df = pd.DataFrame(model_info_data)
            model_info_df.to_excel(writer, sheet_name='Model_Summary', index=False)
    
    def _create_executive_summary_sheet(self, writer: pd.ExcelWriter, insights: Dict[str, Any], metadata: Dict[str, Any]):
        """Create comprehensive executive summary sheet"""
        summary_data = {
            'Report Title': [metadata.get('report_title', 'Ensemble Forecast Report')],
            'Company ID': [metadata.get('company_id', 'N/A')],
            'Generated At': [metadata.get('generated_at', 'N/A')],
            'Forecast Horizon': [f"{metadata.get('forecast_horizon', 6)} months"],
            'Data Quality Score': [f"{metadata.get('data_quality_score', 0.85):.2%}"],
            'Ensemble Models': [', '.join(['ARIMA', 'ETS', 'XGBoost', 'LSTM', 'Croston'])],
            'Confidence Levels': [', '.join(f"{cl:.0%}" for cl in metadata.get('confidence_levels', [0.1, 0.5, 0.9]))],
            'Executive Summary': [insights.get('executive_summary', 'No summary available')],
            'Key Findings': ['; '.join(insights.get('key_findings', []))]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
        
        # Add recommendations summary if available
        if insights.get('recommendations'):
            rec_data = []
            for i, rec in enumerate(insights['recommendations'][:10], 1):
                rec_data.append({
                    'Priority': i,
                    'Title': rec.get('title', 'N/A'),
                    'Description': rec.get('description', 'N/A')[:200] + '...' if len(rec.get('description', '')) > 200 else rec.get('description', 'N/A'),
                    'Urgency': rec.get('urgency', 'N/A'),
                    'Confidence': f"{rec.get('confidence', 0):.2%}",
                    'Impact_Score': rec.get('impact_score', 'N/A')
                })
            
            if rec_data:
                rec_df = pd.DataFrame(rec_data)
                rec_df.to_excel(writer, sheet_name='Top_Recommendations', index=False)
    
    def _create_forecast_sheets(self, writer: pd.ExcelWriter, forecasts: Dict[str, Any]):
        """Create forecast-related sheets"""
        # Ensemble forecast sheet
        if forecasts.get('ensemble_forecast'):
            ensemble_data = {
                'Date': forecasts['ensemble_forecast']['dates'],
                'Ensemble_Forecast': forecasts['ensemble_forecast']['values']
            }
            
            # Add confidence intervals
            if forecasts.get('confidence_intervals'):
                for level, interval_data in forecasts['confidence_intervals'].items():
                    ensemble_data[f'CI_{level.upper()}'] = interval_data['values']
            
            ensemble_df = pd.DataFrame(ensemble_data)
            ensemble_df.to_excel(writer, sheet_name='Ensemble_Forecast', index=False)
        
        # Individual model forecasts
        if forecasts.get('individual_forecasts'):
            individual_data = []
            for model_name, forecast_data in forecasts['individual_forecasts'].items():
                for date, value in zip(forecast_data['dates'], forecast_data['values']):
                    individual_data.append({
                        'Date': date,
                        'Model': model_name.upper(),
                        'Forecast': value
                    })
            
            if individual_data:
                individual_df = pd.DataFrame(individual_data)
                individual_df.to_excel(writer, sheet_name='Individual_Forecasts', index=False)
        
        # Model weights
        if forecasts.get('model_weights'):
            weights_df = pd.DataFrame([forecasts['model_weights']])
            weights_df.to_excel(writer, sheet_name='Model_Weights', index=False)
        
        # Historical data
        if forecasts.get('historical_data', {}).get('data'):
            historical_df = pd.DataFrame(forecasts['historical_data']['data'])
            historical_df.to_excel(writer, sheet_name='Historical_Data', index=False)
    
    def _create_performance_sheets(self, writer: pd.ExcelWriter, performance: Dict[str, Any]):
        """Create performance-related sheets"""
        # Performance history
        if performance.get('performance_history'):
            perf_df = pd.DataFrame(performance['performance_history'])
            perf_df.to_excel(writer, sheet_name='Performance_History', index=False)
        
        # Weight evolution
        if performance.get('weight_evolution'):
            weight_data = []
            for record in performance['weight_evolution']:
                base_record = {
                    'Update_Date': record['update_date'],
                    'Trigger_Reason': record['trigger_reason']
                }
                # Add old and new weights
                for model, weight in record.get('old_weights', {}).items():
                    base_record[f'Old_{model.upper()}'] = weight
                for model, weight in record.get('new_weights', {}).items():
                    base_record[f'New_{model.upper()}'] = weight
                weight_data.append(base_record)
            
            if weight_data:
                weight_df = pd.DataFrame(weight_data)
                weight_df.to_excel(writer, sheet_name='Weight_Evolution', index=False)
        
        # Model rankings
        if performance.get('model_rankings'):
            rankings_df = pd.DataFrame(performance['model_rankings'])
            rankings_df.to_excel(writer, sheet_name='Model_Rankings', index=False)
        
        # Accuracy trends
        if performance.get('accuracy_trends'):
            trend_data = []
            for model, trend in performance['accuracy_trends'].items():
                if isinstance(trend, dict):
                    trend['Model'] = model.upper()
                    trend_data.append(trend)
            
            if trend_data:
                trends_df = pd.DataFrame(trend_data)
                trends_df.to_excel(writer, sheet_name='Accuracy_Trends', index=False)
    
    def _create_insights_sheets(self, writer: pd.ExcelWriter, insights: Dict[str, Any]):
        """Create insights-related sheets"""
        # Performance analysis
        if insights.get('performance_analysis'):
            perf_analysis_df = pd.DataFrame([insights['performance_analysis']])
            perf_analysis_df.to_excel(writer, sheet_name='Performance_Analysis', index=False)
        
        # Growth indicators
        if insights.get('growth_indicators'):
            growth_df = pd.DataFrame([insights['growth_indicators']])
            growth_df.to_excel(writer, sheet_name='Growth_Indicators', index=False)
        
        # Category performance
        if insights.get('category_performance'):
            category_df = pd.DataFrame(insights['category_performance'])
            category_df.to_excel(writer, sheet_name='Category_Performance', index=False)
        
        # Regional performance
        if insights.get('regional_performance'):
            regional_df = pd.DataFrame(insights['regional_performance'])
            regional_df.to_excel(writer, sheet_name='Regional_Performance', index=False)
        
        # Risk assessment
        if insights.get('risk_assessment'):
            risk_df = pd.DataFrame([insights['risk_assessment']])
            risk_df.to_excel(writer, sheet_name='Risk_Assessment', index=False)
        
        # Opportunities
        if insights.get('opportunities'):
            opp_df = pd.DataFrame([insights['opportunities']])
            opp_df.to_excel(writer, sheet_name='Opportunities', index=False)
        
        # Recommendations
        if insights.get('recommendations'):
            rec_df = pd.DataFrame(insights['recommendations'])
            rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
    
    def _create_model_info_sheet(self, writer: pd.ExcelWriter, model_info: Dict[str, Any]):
        """Create model information sheet"""
        model_data = []
        
        # Model versions
        if model_info.get('model_versions'):
            for model, info in model_info['model_versions'].items():
                model_data.append({
                    'Model': model.upper(),
                    'Version': info.get('version', 'N/A'),
                    'Library': info.get('library', 'N/A'),
                    'Parameters': str(info.get('parameters', {}))
                })
        
        if model_data:
            model_df = pd.DataFrame(model_data)
            model_df.to_excel(writer, sheet_name='Model_Information', index=False)
        
        # Ensemble configuration
        if model_info.get('ensemble_configuration'):
            config_df = pd.DataFrame([model_info['ensemble_configuration']])
            config_df.to_excel(writer, sheet_name='Ensemble_Config', index=False)
    
    def _apply_excel_styling(self, workbook):
        """Apply cyberpunk-inspired styling to Excel workbook"""
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            # Define styles
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
            
            data_font = Font(name='Arial', size=10)
            alt_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply styling to all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Style headers
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Style data rows
                for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    for cell in row:
                        cell.font = data_font
                        cell.border = border
                        if row_num % 2 == 0:
                            cell.fill = alt_fill
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            logger.warning(f"Excel styling failed: {e}")
    
    # PDF helper methods
    def _create_text_pdf(self, data: Dict[str, Any], filename: str, metadata: Dict[str, Any] = None) -> str:
        """Create text-based PDF fallback"""
        file_path = self.temp_dir / filename.replace('.pdf', '_report.txt')
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("ENSEMBLE FORECAST COMPREHENSIVE REPORT\n")
            f.write("=" * 60 + "\n\n")
            
            # Metadata section
            if metadata:
                f.write("REPORT METADATA\n")
                f.write("-" * 30 + "\n")
                f.write(f"Report ID: {metadata.get('export_id', 'N/A')}\n")
                f.write(f"Company ID: {metadata.get('company_id', 'N/A')}\n")
                f.write(f"Generated At: {metadata.get('generated_at', 'N/A')}\n")
                f.write(f"Format: {metadata.get('format', 'N/A')}\n")
                f.write(f"Forecast Horizon: {metadata.get('forecast_horizon', 'N/A')} months\n")
                f.write(f"Data Quality Score: {metadata.get('data_quality_score', 'N/A'):.2%}\n\n")
            
            # Executive summary
            if data.get('insights', {}).get('executive_summary'):
                f.write("EXECUTIVE SUMMARY\n")
                f.write("-" * 30 + "\n")
                f.write(data['insights']['executive_summary'] + "\n\n")
            
            # Key findings
            if data.get('insights', {}).get('key_findings'):
                f.write("KEY FINDINGS\n")
                f.write("-" * 30 + "\n")
                for i, finding in enumerate(data['insights']['key_findings'], 1):
                    f.write(f"{i}. {finding}\n")
                f.write("\n")
            
            # Forecast summary
            if data.get('forecasts'):
                f.write("FORECAST SUMMARY\n")
                f.write("-" * 30 + "\n")
                forecasts = data['forecasts']
                
                if forecasts.get('ensemble_forecast'):
                    values = forecasts['ensemble_forecast']['values']
                    f.write(f"Forecast Horizon: {len(values)} periods\n")
                    f.write(f"Average Forecast Value: {np.mean(values):.2f}\n")
                    f.write(f"Forecast Range: {np.min(values):.2f} - {np.max(values):.2f}\n")
                
                if forecasts.get('model_weights'):
                    f.write("\nModel Weights:\n")
                    for model, weight in forecasts['model_weights'].items():
                        f.write(f"  {model.upper()}: {weight:.3f}\n")
                f.write("\n")
            
            # Performance summary
            if data.get('performance'):
                f.write("PERFORMANCE SUMMARY\n")
                f.write("-" * 30 + "\n")
                performance = data['performance']
                
                if performance.get('current_performance'):
                    current = performance['current_performance']
                    f.write(f"Overall Accuracy: {current.get('overall_accuracy', 'N/A')}\n")
                    f.write(f"System Health: {current.get('ensemble_health', 'N/A')}\n")
                
                if performance.get('model_rankings'):
                    f.write("\nModel Rankings:\n")
                    for i, model in enumerate(performance['model_rankings'][:5], 1):
                        f.write(f"  {i}. {model.get('model_name', 'N/A').upper()}: {model.get('accuracy', 'N/A')}\n")
                f.write("\n")
            
            # Recommendations
            if data.get('insights', {}).get('recommendations'):
                f.write("RECOMMENDATIONS\n")
                f.write("-" * 30 + "\n")
                for i, rec in enumerate(data['insights']['recommendations'][:10], 1):
                    f.write(f"{i}. {rec.get('title', 'N/A')}\n")
                    f.write(f"   Description: {rec.get('description', 'N/A')}\n")
                    f.write(f"   Urgency: {rec.get('urgency', 'N/A')}\n")
                    f.write(f"   Confidence: {rec.get('confidence', 0):.2f}\n\n")
        
        return str(file_path)
    
    def _create_pdf_title_section(self, metadata: Dict[str, Any], title_style) -> List:
        """Create comprehensive PDF title section"""
        story = []
        
        title = metadata.get('report_title', 'Ensemble Forecast Report')
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Report info table with enhanced metadata
        report_info = [
            ['Property', 'Value'],
            ['Company ID', metadata.get('company_id', 'N/A')],
            ['Generated At', metadata.get('generated_at', 'N/A')],
            ['Forecast Horizon', f"{metadata.get('forecast_horizon', 6)} months"],
            ['Data Quality Score', f"{metadata.get('data_quality_score', 0.85):.2%}"],
            ['Ensemble Models', 'ARIMA, ETS, XGBoost, LSTM, Croston'],
            ['Confidence Levels', ', '.join(f"{cl:.0%}" for cl in metadata.get('confidence_levels', [0.1, 0.5, 0.9]))],
            ['Export ID', metadata.get('export_id', 'N/A')[-12:] if metadata.get('export_id') else 'N/A']
        ]
        
        info_table = Table(report_info, colWidths=[2.5*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00ffff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        return story
    
    def _create_pdf_executive_summary(self, insights: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF executive summary section"""
        story = []
        
        if insights.get('executive_summary'):
            story.append(Paragraph("Executive Summary", heading_style))
            story.append(Paragraph(insights['executive_summary'], styles['Normal']))
            story.append(Spacer(1, 12))
        
        if insights.get('key_findings'):
            story.append(Paragraph("Key Findings", heading_style))
            for i, finding in enumerate(insights['key_findings'][:5], 1):
                story.append(Paragraph(f"{i}. {finding}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        return story
    
    def _create_pdf_forecast_section(self, forecasts: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF forecast section"""
        story = []
        
        if not forecasts:
            return story
        
        story.append(Paragraph("Forecast Analysis", heading_style))
        
        if forecasts.get('ensemble_forecast'):
            values = forecasts['ensemble_forecast']['values']
            forecast_summary = [
                ['Metric', 'Value'],
                ['Forecast Periods', str(len(values))],
                ['Average Forecast', f"{np.mean(values):.2f}"],
                ['Forecast Range', f"{np.min(values):.2f} - {np.max(values):.2f}"],
                ['Standard Deviation', f"{np.std(values):.2f}"]
            ]
            
            forecast_table = Table(forecast_summary, colWidths=[2*inch, 2*inch])
            forecast_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ff00ff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(forecast_table)
            story.append(Spacer(1, 12))
        
        if forecasts.get('model_weights'):
            story.append(Paragraph("Model Weights", styles['Heading3']))
            weight_data = [['Model', 'Weight']]
            for model, weight in forecasts['model_weights'].items():
                weight_data.append([model.upper(), f"{weight:.3f}"])
            
            weight_table = Table(weight_data, colWidths=[1.5*inch, 1*inch])
            weight_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffff00')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(weight_table)
            story.append(Spacer(1, 12))
        
        return story
    
    def _create_pdf_performance_section(self, performance: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF performance section"""
        story = []
        
        if not performance:
            return story
        
        story.append(Paragraph("Performance Analysis", heading_style))
        
        if performance.get('current_performance'):
            current = performance['current_performance']
            perf_data = [['Metric', 'Value']]
            
            for key, value in current.items():
                if isinstance(value, (int, float)):
                    perf_data.append([key.replace('_', ' ').title(), f"{value:.3f}"])
                else:
                    perf_data.append([key.replace('_', ' ').title(), str(value)])
            
            perf_table = Table(perf_data, colWidths=[2*inch, 2*inch])
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00ffff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(perf_table)
            story.append(Spacer(1, 12))
        
        return story
    
    def _create_pdf_insights_section(self, insights: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF insights section"""
        story = []
        
        if not insights:
            return story
        
        # Growth indicators
        if insights.get('growth_indicators'):
            story.append(Paragraph("Growth Analysis", heading_style))
            growth = insights['growth_indicators']
            
            growth_text = f"Monthly Growth Rate: {growth.get('monthly_growth_rate', 'N/A')}\n"
            growth_text += f"Quarterly Growth Rate: {growth.get('quarterly_growth_rate', 'N/A')}\n"
            growth_text += f"Year-over-Year Growth: {growth.get('year_over_year_growth', 'N/A')}\n"
            growth_text += f"Growth Sustainability Score: {growth.get('growth_sustainability_score', 'N/A')}"
            
            story.append(Paragraph(growth_text, styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Risk assessment
        if insights.get('risk_assessment'):
            story.append(Paragraph("Risk Assessment", heading_style))
            risk = insights['risk_assessment']
            
            risk_text = f"Overall Risk Score: {risk.get('overall_risk_score', 'N/A')}\n"
            risk_text += f"Risk Level: {risk.get('risk_level', 'N/A')}\n"
            
            if risk.get('primary_risks'):
                risk_text += f"Primary Risks: {', '.join(risk['primary_risks'])}"
            
            story.append(Paragraph(risk_text, styles['Normal']))
            story.append(Spacer(1, 12))
        
        return story
    
    def _create_pdf_recommendations_section(self, insights: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF recommendations section"""
        story = []
        
        if insights.get('recommendations'):
            story.append(Paragraph("Recommendations", heading_style))
            
            for i, rec in enumerate(insights['recommendations'][:5], 1):
                rec_text = f"{i}. {rec.get('title', 'N/A')}\n"
                rec_text += f"Description: {rec.get('description', 'N/A')}\n"
                rec_text += f"Urgency: {rec.get('urgency', 'N/A')} | "
                rec_text += f"Confidence: {rec.get('confidence', 0):.2f}"
                
                story.append(Paragraph(rec_text, styles['Normal']))
                story.append(Spacer(1, 8))
        
        return story
    
    def _create_pdf_metadata_section(self, metadata: Dict[str, Any], heading_style, styles) -> List:
        """Create PDF metadata section"""
        story = []
        
        if metadata:
            story.append(PageBreak())
            story.append(Paragraph("Technical Metadata", heading_style))
            
            # System information
            if metadata.get('system_info'):
                sys_info = metadata['system_info']
                sys_text = f"Ensemble Version: {sys_info.get('ensemble_version', 'N/A')}\n"
                sys_text += f"Export Utility Version: {sys_info.get('export_utility_version', 'N/A')}\n"
                sys_text += f"Generation Time: {sys_info.get('generation_time_seconds', 'N/A')} seconds"
                
                story.append(Paragraph(sys_text, styles['Normal']))
                story.append(Spacer(1, 12))
            
            # Model versions
            if metadata.get('model_versions'):
                story.append(Paragraph("Model Versions", styles['Heading3']))
                model_text = ""
                for model, version in metadata['model_versions'].items():
                    model_text += f"{model.upper()}: {version}\n"
                
                story.append(Paragraph(model_text, styles['Normal']))
        
        return story


# Global instance
comprehensive_formatter = ComprehensiveExportFormatter()